# Developed by George
# 14/02/2023

# Script designed to generate samplesheets for Illumina based runs

import sys
import openpyxl
import csv
import os
import zipfile

# loop through all Excel files passed as arguments
for excel_file in sys.argv[1:]:
    # open excel file
    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
    except zipfile.BadZipFile:
        print("Error: Invalid file format for {}. Skipping...".format(excel_file))
        continue

    # mapping of sheet names to suffixes for the output file names
    suffix_map = {
        'TSO500 Combined Loading': '_combined',
        'TSO500 DNA Loading': '_DNA',
        'TSO500 RNA  Loading': '_RNA'
    }

    # check if all required sheets are present in the workbook
    for sheet_name in ['TSO500 Combined Loading', 'TSO500 DNA Loading', 'TSO500 RNA  Loading']:
        if sheet_name not in workbook.sheetnames:
            print("Error: Sheet {} not found in workbook {}. Skipping...".format(sheet_name, excel_file))
            continue

    # loop through the sheets of the workbook
    for sheet_name in ['TSO500 Combined Loading', 'TSO500 DNA Loading', 'TSO500 RNA  Loading']:
        worksheet = workbook[sheet_name]

        # get the ExperimentName value for naming the samplesheets
        file_name = worksheet['B4'].value + suffix_map[sheet_name] + '.csv'

        # get the directory of the source excel file
        output_dir = os.path.dirname(excel_file)

        # open the output CSV file
        with open(os.path.join(output_dir, file_name), 'w', newline='') as csvfile:
            csvwriter = csv.writer(csvfile)

            # loop through the rows and columns of the sheet
            for row in worksheet.iter_rows(values_only=True):
                # check if the row contains only empty cells
                if not all(cell is None or cell == '' for cell in row):
                    # check if the row contains an "N/A" value
                    if "N/A" in row:
                        print("Error: N/A value found in sheet {}. Skipping...".format(sheet_name))
                        continue
                    # write the row to the CSV file
                    csvwriter.writerow(row)

    print("Samplesheets generated: {}.".format(excel_file))

    # create flag.txt in the same directory as excel_file
    flag_file = os.path.join(os.path.dirname(excel_file), "flag.txt")
    with open(flag_file, "w") as f:
        pass
    print("Flag file created for {}.".format(excel_file))

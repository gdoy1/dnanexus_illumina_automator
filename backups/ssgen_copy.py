import sys
import openpyxl
import csv
import os
import zipfile

# Loop through all Excel files passed as arguments
for excel_file in sys.argv[1:]:
    # Open the Excel file
    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
    except zipfile.BadZipFile:
        print("Error: Invalid file format for {}. Skipping...".format(excel_file))
        continue

    # Define a mapping of sheet names to suffixes for the output file names
    suffix_map = {
        'TSO500 Combined Loading': '_combined',
        'TSO500 DNA Loading': '_DNA',
        'TSO500 RNA  Loading': '_RNA'
    }

    # Check if all required sheets are present in the workbook
    for sheet_name in ['TSO500 Combined Loading', 'TSO500 DNA Loading', 'TSO500 RNA  Loading']:
        if sheet_name not in workbook.sheetnames:
            print("Error: Sheet {} not found in workbook {}. Skipping...".format(sheet_name, excel_file))
            continue

    # Loop through the sheets of the workbook
    for sheet_name in ['TSO500 Combined Loading', 'TSO500 DNA Loading', 'TSO500 RNA  Loading']:
        worksheet = workbook[sheet_name]

        # Get the value of cell B4
        file_name = worksheet['B4'].value + suffix_map[sheet_name] + '.csv'

        # Get the directory of the Excel file
        output_dir = os.path.dirname(excel_file)

        # Open the output CSV file in the directory of the Excel file
        with open(os.path.join(output_dir, file_name), 'w', newline='') as csvfile:
            csvwriter = csv.writer(csvfile)

            # Loop through the rows and columns of the sheet
            for row in worksheet.iter_rows(values_only=True):
                # Check if the row contains only empty cells
                if not all(cell is None or cell == '' for cell in row):
                    # Check if the row contains an "N/A" value
                    if "N/A" in row:
                        print("Error: N/A value found in sheet {}. Skipping...".format(sheet_name))
                        continue
                    # Write the row to the CSV file
                    csvwriter.writerow(row)

    print("CSV generation complete for {}.".format(excel_file))

    # Create flag.txt in the same directory as the excel_file
    flag_file = os.path.join(os.path.dirname(excel_file), "flag.txt")
    with open(flag_file, "w") as f:
        pass
    print("Flag file created for {}.".format(excel_file))
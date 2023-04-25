# Developed by George
# 14/02/2023

# Script designed to move samplesheets to corresponding runfolder based on RunParameters.xml

import sys
import openpyxl
import csv
import os
import zipfile
import subprocess
import re
from datetime import datetime

def load_workbook(excel_file):
    """Load workbook and ensure it is valid."""
    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
    except zipfile.BadZipFile:
        print("Error: Invalid file format for {}. Skipping...".format(excel_file))
        return None
    return workbook

def check_hcap_sheet(workbook):
    """Check for Roche run."""
    for sheet_name in workbook.sheetnames:
        if "HCAP" in sheet_name:
            return True
    return False

def write_csv(sheet, output_dir, file_name):
    """Write out to csv file."""
    try:
        with open(os.path.join(output_dir, file_name), "w", newline="") as csvfile:
            csvwriter = csv.writer(csvfile)
            for row in sheet.iter_rows(values_only=True):
                if not all(cell is None or cell == "" for cell in row):
                    if "N/A" in row:
                        print("Error: N/A value found in sheet {}. Skipping...".format(sheet.title))
                        continue
                    formatted_row = []
                    for cell in row:
                        if isinstance(cell, datetime):
                            formatted_row.append(cell.strftime('%d/%m/%Y'))
                        else:
                            formatted_row.append(cell)
                    csvwriter.writerow(formatted_row)
    except FileNotFoundError:
        print("Error: Unable to write to file {}. Skipping...".format(file_name))

def generate_roche_combined_csv(workbook, excel_file):
    """Create the combined csv required for Roche samplesheet."""
    if "Combined Sample sheet" in workbook.sheetnames:
        worksheet = workbook["Combined Sample sheet"]
        raw_value = worksheet["B5"].value
        
        # Use regex to extract the numeric value of the worklist
        match = re.match(r"(\d+)", raw_value)
        if match:
            extracted_value = match.group(1)
        else:
            print(f"Error: Invalid value in B5 ({raw_value}). Skipping...")
            return None

        output_dir = os.path.dirname(excel_file)
        file_name = 'SampleSheet.csv'  # Changed the file_name to 'SampleSheet.csv'
        write_csv(worksheet, output_dir, file_name)
        print("Roche combined csv generated: {}.".format(excel_file))
        return file_name
    else:
        print("Error: Sheet Combined Sample sheet not found in workbook {}. Skipping...".format(excel_file))
        return None

def call_split_somatic_samplesheet_script(combined_csv_path):
    """Call split_somatic_samplesheet within the relevant directory, chdir back to origin."""
    if os.path.isfile(combined_csv_path):
        split_somatic_samplesheet_path = os.path.abspath("split_somatic_samplesheet.py")
        output_dir = os.path.dirname(combined_csv_path)
        original_dir = os.getcwd()

        os.chdir(output_dir)
        subprocess.run(["python3", split_somatic_samplesheet_path, "-s", combined_csv_path])
        os.chdir(original_dir)

def generate_tso500_samplesheets(workbook, excel_file):
    """Generate the tso500 samplesheets with correct formatting."""
    suffix_map = {
        'TSO500 Combined Loading': 'SampleSheet',
        'TSO500 DNA Loading': '_DNA',
        'TSO500 RNA  Loading': '_RNA'
    }

    output_dir = os.path.dirname(excel_file)
    dir_name = os.path.basename(output_dir)
    
    # Check if the directory name contains an underscore
    generate_rna = "_" in dir_name

    for sheet_name in ['TSO500 Combined Loading', 'TSO500 DNA Loading', 'TSO500 RNA  Loading']:
        if sheet_name not in workbook.sheetnames:
            print("Error: Sheet {} not found in workbook {}. Skipping...".format(sheet_name, excel_file))
            continue

        # Only generate RNA samplesheet if directory name contains an underscore
        if not generate_rna and sheet_name == 'TSO500 RNA  Loading':
            print("Skipping RNA samplesheet generation due to no underscore in directory name.")
            continue

        worksheet = workbook[sheet_name]
        if sheet_name == 'TSO500 Combined Loading':
            file_name = suffix_map[sheet_name] + '.csv'
        else:
            file_name = worksheet['B4'].value + suffix_map[sheet_name] + '.csv'
        write_csv(worksheet, output_dir, file_name)

def main():
    excel_file = os.path.abspath(sys.argv[1])
    workbook = load_workbook(excel_file)
    if workbook is None:
        return

    if check_hcap_sheet(workbook):
        combined_csv = generate_roche_combined_csv(workbook, excel_file)
        if combined_csv is not None:
            call_split_somatic_samplesheet_script(os.path.join(os.path.dirname(excel_file), combined_csv))
    else:
        generate_tso500_samplesheets(workbook, excel_file)
        print("Samplesheets generated: {}.".format(excel_file))

if __name__ == "__main__":
    main()